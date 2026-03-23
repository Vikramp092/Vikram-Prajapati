#!/usr/bin/env python3
"""
Professional Depreciation Calculator
===================================

A comprehensive Python script that generates an Excel workbook for depreciation
calculations under Indian tax laws:

- Companies Act, 2013 (Schedule II - Straight Line Method)
- Income Tax Act, 1961 (Written Down Value Method)

Author: Chartered Accountant & Python Developer
License: MIT
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

class DepreciationCalculator:
    """
    Professional Depreciation Calculator for Indian Tax Compliance

    Generates Excel workbook with dual depreciation calculations:
    - Companies Act, 2013 (SLM Method)
    - Income Tax Act, 1961 (WDV Method)
    """

    def __init__(self):
        """Initialize the calculator with styling and configuration"""
        self.setup_styles()

    def setup_styles(self):
        """Define professional Excel styling"""
        # Header style - Professional blue header
        self.header_style = NamedStyle(
            name="header",
            font=Font(bold=True, size=12, color="FFFFFF"),
            fill=PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
        )

        # Data style - Clean bordered cells
        self.data_style = NamedStyle(
            name="data",
            alignment=Alignment(horizontal="left", vertical="center"),
            border=Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        )

        # Currency style - Indian Rupee formatting
        self.currency_style = NamedStyle(
            name="currency",
            number_format='"₹"#,##0.00',
            alignment=Alignment(horizontal="right", vertical="center"),
            border=Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        )

        # Percentage style
        self.percent_style = NamedStyle(
            name="percent",
            number_format='0.00%',
            alignment=Alignment(horizontal="center", vertical="center"),
            border=Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        )

        # Date style
        self.date_style = NamedStyle(
            name="date",
            number_format='DD/MM/YYYY',
            alignment=Alignment(horizontal="center", vertical="center"),
            border=Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        )

        # Total row style - Bold with light background
        self.total_style = NamedStyle(
            name="total",
            font=Font(bold=True, size=11),
            fill=PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
        )

    def create_input_sheet(self, wb):
        """
        Create the Input sheet with asset data entry

        Args:
            wb: openpyxl Workbook object
        """
        sheet = wb.create_sheet("Input")

        # Define headers as per requirements
        headers = [
            "Asset Name", "Asset Category", "Date of Purchase", "Date Put to Use",
            "Cost of Asset", "Residual Value", "Useful Life (Schedule II)",
            "Depreciation Rate (Income Tax)", "Block of Asset", "Opening WDV",
            "Additions", "Deletions"
        ]

        # Apply headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = "header"

        # Sample data for demonstration (Indian context)
        sample_data = [
            ["Office Computer", "Plant & Machinery", "01/04/2023", "01/04/2023", 50000, 5000, 5, 15.0, "Block A", 0, 0, 0],
            ["Office Furniture", "Furniture & Fittings", "15/06/2023", "15/06/2023", 25000, 2500, 10, 10.0, "Block B", 0, 0, 0],
            ["Factory Machine", "Plant & Machinery", "01/01/2023", "01/07/2023", 200000, 20000, 15, 7.5, "Block A", 0, 0, 0],
            ["Building", "Building", "01/04/2022", "01/04/2022", 1000000, 100000, 60, 5.0, "Block C", 0, 0, 0],
            ["Motor Vehicle", "Motor Cars", "01/10/2023", "01/10/2023", 800000, 80000, 8, 15.0, "Block D", 0, 0, 0],
            ["Server Equipment", "Plant & Machinery", "01/07/2023", "15/08/2023", 150000, 15000, 6, 15.0, "Block A", 0, 0, 0]
        ]

        # Apply sample data with appropriate styling
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)

                # Apply appropriate styling based on column type
                if col in [5, 6, 10, 11, 12]:  # Currency columns
                    cell.style = "currency"
                elif col == 8:  # Percentage
                    cell.style = "percent"
                elif col in [3, 4]:  # Dates
                    cell.style = "date"
                else:
                    cell.style = "data"

        # Add data validation for Asset Category dropdown
        category_validation = DataValidation(
            type="list",
            formula1='"Plant & Machinery,Furniture & Fittings,Building,Motor Cars,Office Equipment,Computers,Server Equipment"',
            allow_blank=True
        )
        category_validation.error = "Please select a valid asset category"
        category_validation.errorTitle = "Invalid Category"
        sheet.add_data_validation(category_validation)
        category_validation.add('B2:B1000')  # Apply to column B, rows 2-1000

        # Add data validation for Block of Asset dropdown
        block_validation = DataValidation(
            type="list",
            formula1='"Block A,Block B,Block C,Block D"',
            allow_blank=True
        )
        block_validation.error = "Please select a valid block"
        block_validation.errorTitle = "Invalid Block"
        sheet.add_data_validation(block_validation)
        block_validation.add('I2:I1000')  # Apply to column I, rows 2-1000

        # Set column widths for better readability
        column_widths = [25, 20, 15, 15, 15, 15, 20, 25, 12, 15, 12, 12]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        return sheet

    def create_companies_act_sheet(self, wb):
        """
        Create Companies Act depreciation sheet (SLM Method)

        Args:
            wb: openpyxl Workbook object
        """
        sheet = wb.create_sheet("Companies Act Depreciation")

        # Define headers for Companies Act calculations
        headers = [
            "Asset Name", "Cost of Asset", "Residual Value", "Useful Life",
            "Date of Purchase", "Date Put to Use", "Days Used in Year",
            "Annual Depreciation Rate", "Depreciation for Year",
            "Accumulated Depreciation", "Closing WDV"
        ]

        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = "header"

        # Create formulas for each asset (rows 2-7 for 6 sample assets)
        for row in range(2, 8):
            input_row = row  # Corresponding row in Input sheet

            # Asset Name (reference from Input sheet)
            sheet.cell(row=row, column=1, value=f"=Input!A{input_row}").style = "data"

            # Cost of Asset
            sheet.cell(row=row, column=2, value=f"=Input!E{input_row}").style = "currency"

            # Residual Value
            sheet.cell(row=row, column=3, value=f"=Input!F{input_row}").style = "currency"

            # Useful Life
            sheet.cell(row=row, column=4, value=f"=Input!G{input_row}").style = "data"

            # Date of Purchase
            sheet.cell(row=row, column=5, value=f"=Input!C{input_row}").style = "date"

            # Date Put to Use
            sheet.cell(row=row, column=6, value=f"=Input!D{input_row}").style = "date"

            # Days Used in Year - Calculate based on financial year (April-March)
            # Formula: If put to use before April, use 365, else calculate days from put to use to March 31
            sheet.cell(row=row, column=7,
                value=f"=IF(MONTH(Input!D{input_row})<4,365,DAYS(DATE(YEAR(Input!D{input_row})+1,3,31),Input!D{input_row}))"
            ).style = "data"

            # Annual Depreciation Rate (SLM) = (Cost - Residual) / Useful Life
            sheet.cell(row=row, column=8,
                value=f"=(B{row}-C{row})/D{row}"
            ).style = "currency"

            # Depreciation for Year - Pro-rata if used <180 days
            # Formula: If days used <180, then Annual Rate * (Days/365), else Annual Rate
            sheet.cell(row=row, column=9,
                value=f"=IF(G{row}<180,H{row}*G{row}/365,H{row})"
            ).style = "currency"

            # Accumulated Depreciation (starts with current year depreciation)
            sheet.cell(row=row, column=10, value=f"=I{row}").style = "currency"

            # Closing WDV = Cost - Accumulated Depreciation
            sheet.cell(row=row, column=11, value=f"=B{row}-J{row}").style = "currency"

        # Add totals row
        sheet.cell(row=8, column=1, value="TOTALS").style = "total"
        sheet.cell(row=8, column=2, value="=SUM(B2:B7)").style = "currency"
        sheet.cell(row=8, column=3, value="=SUM(C2:C7)").style = "currency"
        sheet.cell(row=8, column=9, value="=SUM(I2:I7)").style = "currency"
        sheet.cell(row=8, column=10, value="=SUM(J2:J7)").style = "currency"
        sheet.cell(row=8, column=11, value="=SUM(K2:K7)").style = "currency"

        # Set column widths
        column_widths = [25, 15, 15, 12, 15, 15, 15, 20, 20, 20, 15]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        return sheet

    def create_income_tax_sheet(self, wb):
        """
        Create Income Tax Act depreciation sheet (WDV Method)

        Args:
            wb: openpyxl Workbook object
        """
        sheet = wb.create_sheet("Income Tax Depreciation")

        # Define headers for Income Tax calculations
        headers = [
            "Block of Asset", "Opening WDV", "Additions During Year",
            "Deletions During Year", "Total WDV", "Depreciation Rate",
            "Normal Depreciation", "Additional Depreciation", "Total Depreciation",
            "Closing WDV", "Assets in Block"
        ]

        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = "header"

        # Group assets by block for WDV calculations
        # Block A: Plant & Machinery (15% rate)
        # Block B: Furniture & Fittings (10% rate)
        # Block C: Building (5% rate)
        # Block D: Motor Cars (15% rate)

        block_data = {
            "Block A": {"rate": 0.15, "assets": [1, 3, 6]},  # Office Computer, Factory Machine, Server Equipment
            "Block B": {"rate": 0.10, "assets": [2]},        # Office Furniture
            "Block C": {"rate": 0.05, "assets": [4]},        # Building
            "Block D": {"rate": 0.15, "assets": [5]}         # Motor Vehicle
        }

        row = 2
        for block_name, block_info in block_data.items():
            # Block of Asset
            sheet.cell(row=row, column=1, value=block_name).style = "data"

            # Opening WDV (from Input sheet)
            opening_refs = [f"Input!J{asset_row}" for asset_row in block_info["assets"]]
            sheet.cell(row=row, column=2, value=f"=SUM({','.join(opening_refs)})").style = "currency"

            # Additions During Year (Cost of assets in block)
            addition_refs = [f"Input!E{asset_row}" for asset_row in block_info["assets"]]
            sheet.cell(row=row, column=3, value=f"=SUM({','.join(addition_refs)})").style = "currency"

            # Deletions During Year
            deletion_refs = [f"Input!L{asset_row}" for asset_row in block_info["assets"]]
            sheet.cell(row=row, column=4, value=f"=SUM({','.join(deletion_refs)})").style = "currency"

            # Total WDV = Opening + Additions - Deletions
            sheet.cell(row=row, column=5, value=f"=B{row}+C{row}-D{row}").style = "currency"

            # Depreciation Rate
            sheet.cell(row=row, column=6, value=block_info["rate"]).style = "percent"

            # Normal Depreciation = Total WDV × Rate
            sheet.cell(row=row, column=7, value=f"=E{row}*F{row}").style = "currency"

            # Additional Depreciation - 50% of normal if any asset used <180 days
            # Check if any asset in block was used <180 days
            days_checks = []
            for asset_row in block_info["assets"]:
                # Days used calculation (same as Companies Act sheet)
                days_formula = f"IF(MONTH(Input!D{asset_row})<4,365,DAYS(DATE(YEAR(Input!D{asset_row})+1,3,31),Input!D{asset_row}))"
                days_checks.append(f"({days_formula}<180)")

            additional_check = "OR(" + ",".join(days_checks) + ")"
            sheet.cell(row=row, column=8, value=f"=IF({additional_check},G{row}*0.5,0)").style = "currency"

            # Total Depreciation = Normal + Additional
            sheet.cell(row=row, column=9, value=f"=G{row}+H{row}").style = "currency"

            # Closing WDV = Total WDV - Total Depreciation
            sheet.cell(row=row, column=10, value=f"=E{row}-I{row}").style = "currency"

            # Assets in Block (count)
            sheet.cell(row=row, column=11, value=len(block_info["assets"])).style = "data"

            row += 1

        # Add totals row
        sheet.cell(row=row, column=1, value="TOTALS").style = "total"
        sheet.cell(row=row, column=2, value=f"=SUM(B2:B{row-1})").style = "currency"
        sheet.cell(row=row, column=3, value=f"=SUM(C2:C{row-1})").style = "currency"
        sheet.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})").style = "currency"
        sheet.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").style = "currency"
        sheet.cell(row=row, column=7, value=f"=SUM(G2:G{row-1})").style = "currency"
        sheet.cell(row=row, column=8, value=f"=SUM(H2:H{row-1})").style = "currency"
        sheet.cell(row=row, column=9, value=f"=SUM(I2:I{row-1})").style = "currency"
        sheet.cell(row=row, column=10, value=f"=SUM(J2:J{row-1})").style = "currency"

        # Set column widths
        column_widths = [15, 15, 18, 18, 15, 15, 18, 20, 18, 15, 15]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        return sheet

    def create_summary_sheet(self, wb):
        """
        Create summary sheet with totals and variance analysis

        Args:
            wb: openpyxl Workbook object
        """
        sheet = wb.create_sheet("Summary")

        # Title
        sheet.cell(row=1, column=1, value="DEPRECIATION CALCULATOR - SUMMARY & ANALYSIS").style = "header"
        sheet.merge_cells('A1:G1')

        # Companies Act Summary
        sheet.cell(row=3, column=1, value="COMPANIES ACT, 2013 (SLM METHOD)").style = "header"
        sheet.merge_cells('A3:G3')

        ca_summaries = [
            ("Total Cost of Assets", "=SUM('Companies Act Depreciation'!B2:B7)"),
            ("Total Residual Value", "=SUM('Companies Act Depreciation'!C2:C7)"),
            ("Total Depreciation for Year", "=SUM('Companies Act Depreciation'!I2:I7)"),
            ("Total Closing WDV", "=SUM('Companies Act Depreciation'!K2:K7)")
        ]

        for i, (label, formula) in enumerate(ca_summaries, 4):
            sheet.cell(row=i, column=1, value=label).style = "data"
            sheet.cell(row=i, column=2, value="=").style = "data"
            sheet.cell(row=i, column=3, value=formula).style = "currency"

        # Income Tax Act Summary
        sheet.cell(row=9, column=1, value="INCOME TAX ACT, 1961 (WDV METHOD)").style = "header"
        sheet.merge_cells('A9:G9')

        it_summaries = [
            ("Total Opening WDV", "=SUM('Income Tax Depreciation'!B2:B5)"),
            ("Total Additions", "=SUM('Income Tax Depreciation'!C2:C5)"),
            ("Total Deletions", "=SUM('Income Tax Depreciation'!D2:D5)"),
            ("Total Normal Depreciation", "=SUM('Income Tax Depreciation'!G2:G5)"),
            ("Total Additional Depreciation", "=SUM('Income Tax Depreciation'!H2:H5)"),
            ("Total Depreciation", "=SUM('Income Tax Depreciation'!I2:I5)"),
            ("Total Closing WDV", "=SUM('Income Tax Depreciation'!J2:J5)")
        ]

        for i, (label, formula) in enumerate(it_summaries, 10):
            sheet.cell(row=i, column=1, value=label).style = "data"
            sheet.cell(row=i, column=2, value="=").style = "data"
            sheet.cell(row=i, column=3, value=formula).style = "currency"

        # Variance Analysis
        sheet.cell(row=18, column=1, value="VARIANCE ANALYSIS").style = "header"
        sheet.merge_cells('A18:G18')

        variances = [
            ("Book vs Tax Depreciation", "=C7-C16"),
            ("Book vs Tax WDV", "=C8-C17"),
            ("Additional Depreciation Impact", "=C15")
        ]

        for i, (label, formula) in enumerate(variances, 19):
            sheet.cell(row=i, column=1, value=label).style = "data"
            sheet.cell(row=i, column=2, value="=").style = "data"
            sheet.cell(row=i, column=3, value=formula).style = "currency"

        # Instructions
        sheet.cell(row=23, column=1, value="HOW TO USE THIS CALCULATOR").style = "header"
        sheet.merge_cells('A23:G23')

        instructions = [
            "1. Enter asset details in the 'Input' sheet",
            "2. Select appropriate categories and blocks from dropdowns",
            "3. Review Companies Act calculations (SLM with pro-rata)",
            "4. Review Income Tax calculations (WDV with block concept)",
            "5. Use this summary for variance analysis and reporting",
            "6. All calculations update automatically when input changes",
            "7. Save regularly and backup your depreciation working papers"
        ]

        for i, instruction in enumerate(instructions, 24):
            sheet.cell(row=i, column=1, value=instruction).style = "data"
            sheet.merge_cells(f'A{i}:G{i}')

        # Set column widths
        column_widths = [30, 5, 20, 5, 5, 5, 5]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        return sheet

    def generate_calculator(self, output_file="Depreciation_Calculator.xlsx"):
        """
        Generate the complete depreciation calculator Excel file

        Args:
            output_file (str): Name of the output Excel file
        """
        print("🏗️  Building Professional Depreciation Calculator...")
        print("=" * 60)

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Apply named styles to workbook
        wb.add_named_style(self.header_style)
        wb.add_named_style(self.data_style)
        wb.add_named_style(self.currency_style)
        wb.add_named_style(self.percent_style)
        wb.add_named_style(self.date_style)
        wb.add_named_style(self.total_style)

        # Create all sheets
        print("📝 Creating Input sheet...")
        self.create_input_sheet(wb)

        print("🏢 Creating Companies Act depreciation sheet...")
        self.create_companies_act_sheet(wb)

        print("💰 Creating Income Tax depreciation sheet...")
        self.create_income_tax_sheet(wb)

        print("📊 Creating Summary & Analysis sheet...")
        self.create_summary_sheet(wb)

        # Save the workbook
        print(f"💾 Saving to {output_file}...")
        wb.save(output_file)

        print("\n✅ SUCCESS! Professional Depreciation Calculator created!")
        print("=" * 60)
        print(f"📁 File: {output_file}")
        print(f"📊 Size: {os.path.getsize(output_file)} bytes")
        print("\n📋 SHEETS CREATED:")
        print("• Input - Asset data entry with validation")
        print("• Companies Act Depreciation - SLM calculations")
        print("• Income Tax Depreciation - WDV calculations")
        print("• Summary - Analysis and variance reporting")

        print("\n🎯 KEY FEATURES:")
        print("• Automatic pro-rata depreciation (<180 days)")
        print("• Block-wise WDV calculations")
        print("• Professional Excel formatting")
        print("• Data validation and dropdowns")
        print("• Dynamic formulas (not static values)")
        print("• Variance analysis between methods")
        print("• Suitable for audit and tax compliance")

        return output_file

def main():
    """
    Main function to run the depreciation calculator generator
    """
    print("🇮🇳 Professional Depreciation Calculator")
    print("Chartered Accountant & Python Developer")
    print("=" * 60)

    # Create calculator instance
    calculator = DepreciationCalculator()

    # Generate the Excel file
    output_file = calculator.generate_calculator()

    print(f"\n🚀 Ready to use! Open '{output_file}' in Excel to start calculating depreciation.")
    print("\n📚 INSTRUCTIONS:")
    print("1. Open the Excel file")
    print("2. Go to 'Input' sheet and enter your asset details")
    print("3. Review calculations in 'Companies Act Depreciation' and 'Income Tax Depreciation' sheets")
    print("4. Use 'Summary' sheet for variance analysis and reporting")

    return output_file

if __name__ == "__main__":
    main()